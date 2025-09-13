import sys
import cv2
import numpy as np
import pandas as pd
from PyQt5.QtWidgets import (
    QApplication, QLabel, QPushButton, QVBoxLayout, QWidget, QFileDialog,
    QComboBox, QListWidget, QListWidgetItem, QMessageBox, QHBoxLayout, QInputDialog, QMenu,
    QTableWidget, QTableWidgetItem, QDialog, QHeaderView, QVBoxLayout as QVBoxDialogLayout
)
from PyQt5.QtGui import QPixmap, QImage, QColor, QIcon
from PyQt5.QtCore import Qt, QPoint, QSize
from scipy.optimize import nnls
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# 全局变量：存储选取的颜色
picked_colors = []

# ---------------- 工具函数 ----------------
def normalize_rgb(rgb):
    """将RGB值归一化到[0, 1]范围"""
    return np.array(rgb) / 255.0

def analyze_color_mix(base_colors, target_colors, threshold=0.05):
    """分析目标颜色是否可以通过基础颜色混合得到"""
    results = []
    if not base_colors or not target_colors:
        return results

    base_arr = np.array([normalize_rgb(c["RGB"]) for c in base_colors])
    base_hex = [c["HEX"] for c in base_colors]

    for target in target_colors:
        target_rgb = normalize_rgb(target["RGB"])
        alpha, rnorm = nnls(base_arr.T, target_rgb)
        fitted = np.dot(alpha, base_arr)
        error = np.linalg.norm(fitted - target_rgb)

        contributing = [(base_hex[i], alpha[i]) for i in range(len(base_arr)) if alpha[i] > 1e-3]

        results.append({
            "target_color": target["HEX"],
            "target_label": target.get("CustomLabel", target["Label"]),
            "target_rgb": target["RGB"],  # 存储目标色的RGB用于背景色
            "success": error <= threshold,
            "error": error,
            "alpha": alpha,  # 存储完整的alpha数组
            "contributing_colors": contributing
        })
    return results

# ---------------- UI界面 ----------------
class ColorPickerApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("颜色取样与混合分析")
        self.resize(900, 800)

        # 图片显示区域
        self.image_label = QLabel("请加载一张图片")
        self.image_label.setAlignment(Qt.AlignCenter)
        self.image_label.setStyleSheet("background: #ddd;")
        self.image_label.setFixedHeight(400)

        # 标签选择
        self.combo_label = QLabel("选择颜色类型：")
        self.combo_box = QComboBox()
        self.combo_box.addItems(["基础色", "目标色"])

        # 控制按钮
        self.load_button = QPushButton("加载图片")
        self.analyze_button = QPushButton("运行颜色分析")
        self.clear_button = QPushButton("清空颜色列表")
        self.import_labels_button = QPushButton("导入标签")  # 新增导入标签按钮

        control_layout = QHBoxLayout()
        control_layout.addWidget(self.combo_label)
        control_layout.addWidget(self.combo_box)
        control_layout.addWidget(self.load_button)
        control_layout.addWidget(self.analyze_button)
        control_layout.addWidget(self.clear_button)
        control_layout.addWidget(self.import_labels_button)

        # 基础色列表
        self.base_label = QLabel("基础色列表：")
        self.base_list = QListWidget()
        self.base_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.base_list.customContextMenuRequested.connect(lambda pos: self.show_context_menu(pos, self.base_list))
        self.base_list.itemDoubleClicked.connect(self.edit_label)

        # 目标色列表
        self.target_label = QLabel("目标色列表：")
        self.target_list = QListWidget()
        self.target_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.target_list.customContextMenuRequested.connect(lambda pos: self.show_context_menu(pos, self.target_list))
        self.target_list.itemDoubleClicked.connect(self.edit_label)

        # 主布局
        layout = QVBoxLayout()
        layout.addWidget(self.image_label)
        layout.addLayout(control_layout)
        layout.addWidget(self.base_label)
        layout.addWidget(self.base_list)
        layout.addWidget(self.target_label)
        layout.addWidget(self.target_list)
        self.setLayout(layout)

        # 绑定事件
        self.load_button.clicked.connect(self.load_image)
        self.analyze_button.clicked.connect(self.run_analysis)
        self.clear_button.clicked.connect(self.clear_colors)
        self.import_labels_button.clicked.connect(self.import_labels)  # 绑定导入标签按钮

        self.img = None
        self.qimg = None
        self.display_pixmap = None

    def refresh_lists(self):
        """刷新基础色和目标色列表显示，添加颜色正方形图标"""
        self.base_list.clear()
        self.target_list.clear()
        for i, color in enumerate(picked_colors):
            display_label = color.get("CustomLabel", color["Label"])
            item_text = f"{display_label}: {color['HEX']} RGB{color['RGB']}"
            item = QListWidgetItem(item_text)
            item.setData(Qt.UserRole, i)  # 存储全局索引

            # 创建颜色正方形图标
            pixmap = QPixmap(16, 16)
            pixmap.fill(QColor(*color["RGB"]))
            item.setIcon(QIcon(pixmap))

            if color["Label"] == "基础色":
                self.base_list.addItem(item)
            else:
                self.target_list.addItem(item)

    def load_image(self):
        """加载图片并显示"""
        file_path, _ = QFileDialog.getOpenFileName(self, "选择图片", "", "图片文件 (*.png *.jpg *.jpeg)")
        if file_path:
            self.img = cv2.imread(file_path)
            if self.img is None:
                QMessageBox.warning(self, "错误", "无法加载图片，请检查文件格式！")
                return
            rgb = cv2.cvtColor(self.img, cv2.COLOR_BGR2RGB)
            h, w, ch = rgb.shape
            self.qimg = QImage(rgb.data, w, h, ch * w, QImage.Format_RGB888).copy()
            self.display_pixmap = QPixmap.fromImage(self.qimg)

            self.image_label.setPixmap(self.display_pixmap.scaled(
                self.image_label.width(), self.image_label.height(), Qt.KeepAspectRatio, Qt.SmoothTransformation
            ))
            self.image_label.mousePressEvent = self.get_pixel_color

    def get_pixel_color(self, event):
        """获取点击位置的像素颜色，不提示输入标签"""
        if self.img is None or self.display_pixmap is None:
            return

        pixmap = self.image_label.pixmap()
        if not pixmap:
            return

        # 计算点击坐标到pixmap的映射
        label_w, label_h = self.image_label.width(), self.image_label.height()
        pixmap_w, pixmap_h = pixmap.width(), pixmap.height()

        x_offset = (label_w - pixmap_w) / 2
        y_offset = (label_h - pixmap_h) / 2

        x = event.pos().x() - x_offset
        y = event.pos().y() - y_offset

        if x < 0 or y < 0 or x >= pixmap_w or y >= pixmap_h:
            return  # 点击在空白区域

        # 映射到原图坐标（提高精度）
        img_h, img_w, _ = self.img.shape
        img_x = int(x * img_w / pixmap_w)
        img_y = int(y * img_h / pixmap_h)

        # 确保坐标在有效范围内
        img_x = min(max(img_x, 0), img_w - 1)
        img_y = min(max(img_y, 0), img_h - 1)

        b, g, r = self.img[img_y, img_x].tolist()
        hex_code = "#{:02x}{:02x}{:02x}".format(r, g, b)
        label = self.combo_box.currentText()

        color_dict = {"x": img_x, "y": img_y, "RGB": (r, g, b), "HEX": hex_code, "Label": label}
        picked_colors.append(color_dict)

        self.refresh_lists()

    def import_labels(self):
        """批量导入标签，格式为逗号分隔的字符串"""
        text, ok = QInputDialog.getText(self, "导入标签", "请输入标签（用逗号分隔，例如：1,2,3,4）：")
        if not ok or not text.strip():
            return

        # 解析输入的标签
        labels = [label.strip() for label in text.split(",") if label.strip()]
        if not labels:
            QMessageBox.warning(self, "提示", "输入的标签为空或无效！")
            return

        # 按顺序分配标签
        for i, label in enumerate(labels):
            if i < len(picked_colors):
                picked_colors[i]["CustomLabel"] = label

        self.refresh_lists()
        QMessageBox.information(self, "导入成功", f"已为 {min(len(labels), len(picked_colors))} 个颜色分配标签")

    def create_analysis_table(self, results, base_colors):
        """创建并显示颜色混合分析表格，同时返回表格数据用于导出"""
        if not results:
            return None

        # 基础色信息
        base_labels = [c.get("CustomLabel", c["Label"]) for c in base_colors]
        base_hex = [c["HEX"] for c in base_colors]
        base_rgb = [c["RGB"] for c in base_colors]

        # 表格行数：目标色数量 + 1（表头）
        num_rows = len(results) + 1
        # 表格列数：基础色数量 + 1（目标色）
        num_cols = len(base_colors) + 1

        # 创建对话框和表格
        dialog = QDialog(self)
        dialog.setWindowTitle("颜色混合分析结果")
        dialog.resize(800, 400)
        layout = QVBoxDialogLayout()
        table = QTableWidget()
        table.setRowCount(num_rows)
        table.setColumnCount(num_cols)
        layout.addWidget(table)
        dialog.setLayout(layout)

        # 设置表头并添加背景色
        header_labels = ["目标色"] + [f"{label}({hex_code})" for label, hex_code in zip(base_labels, base_hex)]
        table.setHorizontalHeaderLabels(header_labels)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        for col_idx in range(num_cols):
            header_item = table.horizontalHeaderItem(col_idx)
            if header_item and col_idx > 0:  # 跳过“目标色”列
                rgb_color = base_rgb[col_idx - 1]
                header_item.setBackground(QColor(rgb_color[0], rgb_color[1], rgb_color[2]))

        # 创建 Excel 数据
        excel_data = {
            "目标色": [],
            **{f"{label}({hex_code})": [] for label, hex_code in zip(base_labels, base_hex)}
        }

        # 填充表格和 Excel 数据
        for row_idx, result in enumerate(results):
            # 第一列：目标色 HEX，设置背景色
            target_item = QTableWidgetItem(result["target_color"])
            target_rgb = result["target_rgb"]
            target_item.setBackground(QColor(target_rgb[0], target_rgb[1], target_rgb[2]))
            table.setItem(row_idx + 1, 0, target_item)
            excel_data["目标色"].append(result["target_color"])

            # 填充占比列（无背景色）
            for col_idx, alpha_val in enumerate(result["alpha"]):
                if alpha_val > 1e-3:
                    value_item = QTableWidgetItem(f"{alpha_val:.3f}")
                    excel_data[f"{base_labels[col_idx]}({base_hex[col_idx]})"].append(f"{alpha_val:.3f}")
                else:
                    value_item = QTableWidgetItem("")
                    excel_data[f"{base_labels[col_idx]}({base_hex[col_idx]})"].append("")
                table.setItem(row_idx + 1, col_idx + 1, value_item)

        table.resizeColumnsToContents()
        dialog.exec_()

        # 返回 Excel 数据和颜色信息
        return excel_data, base_rgb, [r["target_rgb"] for r in results]

    def run_analysis(self):
        """运行颜色混合分析，显示表格并导出结果到Excel"""
        base_colors = [c for c in picked_colors if c["Label"] == "基础色"]
        target_colors = [c for c in picked_colors if c["Label"] == "目标色"]

        if not base_colors or not target_colors:
            QMessageBox.warning(self, "提示", "请至少选择一种基础色和一种目标色！")
            return

        results = analyze_color_mix(base_colors, target_colors)

        # 显示表格并获取 Excel 数据
        excel_data, base_rgb, target_rgb = self.create_analysis_table(results, base_colors)

        # 导出分析结果到 Excel
        if excel_data:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Color Analysis Results"

                # 写入表头
                headers = list(excel_data.keys())
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = header
                    if col_idx > 1:  # 为基础色表头设置背景色
                        rgb = base_rgb[col_idx - 2]
                        fill = PatternFill(start_color=f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}", end_color=f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}", fill_type="solid")
                        cell.fill = fill

                # 写入数据
                for row_idx in range(len(excel_data["目标色"])):
                    # 第一列：目标色 HEX，设置背景色
                    cell = ws.cell(row=row_idx + 2, column=1)
                    cell.value = excel_data["目标色"][row_idx]
                    rgb = target_rgb[row_idx]
                    fill = PatternFill(start_color=f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}", end_color=f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}", fill_type="solid")
                    cell.fill = fill

                    # 其他列：占比值，无背景色
                    for col_idx, header in enumerate(headers[1:], 2):
                        cell = ws.cell(row=row_idx + 2, column=col_idx)
                        cell.value = excel_data[header][row_idx]

                wb.save("color_analysis_results.xlsx")
                QMessageBox.information(self, "导出成功", "分析结果已保存至 color_analysis_results.xlsx")
            except Exception as e:
                QMessageBox.warning(self, "导出失败", f"保存分析结果时出错：{str(e)}")

    def clear_colors(self):
        """清空已选取的颜色列表"""
        global picked_colors
        picked_colors = []
        self.refresh_lists()

    def show_context_menu(self, pos: QPoint, list_widget: QListWidget):
        """显示右键菜单，支持基础色和目标色修改标签及删除颜色"""
        item = list_widget.itemAt(pos)
        if item is None:
            return
        menu = QMenu()
        delete_action = menu.addAction("删除颜色")
        edit_action = menu.addAction("修改标签")
        action = menu.exec_(list_widget.mapToGlobal(pos))
        if action == delete_action:
            idx = item.data(Qt.UserRole)
            if idx is not None:
                picked_colors.pop(idx)
                self.refresh_lists()
        elif action == edit_action:
            self.edit_label(item)

    def edit_label(self, item: QListWidgetItem):
        """编辑颜色标签"""
        idx = item.data(Qt.UserRole)
        if idx is None:
            return
        color = picked_colors[idx]
        text, ok = QInputDialog.getText(self, "修改标签", "请输入新标签:", text=color.get("CustomLabel", color["Label"]))
        if ok and text.strip():
            picked_colors[idx]["CustomLabel"] = text.strip()
            self.refresh_lists()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = ColorPickerApp()
    win.show()
    sys.exit(app.exec_())
